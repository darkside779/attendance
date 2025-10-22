"""
Report generation service for PDF and Excel exports
"""
import io
import pandas as pd
from datetime import datetime, date, timedelta
from typing import List, Optional
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from sqlalchemy.orm import Session

from app.models.attendance import Attendance
from app.models.employee import Employee
from app.models.salary import Salary
from app.models.shift import Shift
from app.services.employee_service import EmployeeService
from sqlalchemy import func
import json

class ReportService:
    def __init__(self, db: Session):
        self.db = db
        self.employee_service = EmployeeService(db)
    
    def _calculate_monthly_hours_from_shift(self, employee_id: int) -> float:
        """Calculate monthly hours based on employee's assigned shift"""
        # Get the employee's active shift
        shift = self.db.query(Shift).filter(
            Shift.employee_id == employee_id,
            Shift.is_active == True
        ).first()
        
        if not shift:
            print(f"  No shift found for employee {employee_id}, using default 160 hours")
            return 160.0  # Default fallback
        
        # Calculate daily hours
        start_time = shift.start_time
        end_time = shift.end_time
        
        # Handle shifts that cross midnight (e.g., 16:00-00:00)
        if end_time < start_time:
            # Shift crosses midnight
            daily_hours = (24 - start_time.hour - start_time.minute/60) + (end_time.hour + end_time.minute/60)
        else:
            # Normal shift within same day
            daily_hours = (end_time.hour + end_time.minute/60) - (start_time.hour + start_time.minute/60)
        
        # Parse days of week from JSON
        try:
            days_of_week = json.loads(shift.days_of_week) if isinstance(shift.days_of_week, str) else shift.days_of_week
            days_per_week = len(days_of_week)
        except:
            days_per_week = 5  # Default to 5 days if parsing fails
        
        # Calculate monthly hours (4.33 weeks per month average)
        weekly_hours = daily_hours * days_per_week
        monthly_hours = weekly_hours * 4.33
        
        print(f"  Shift: {shift.shift_name} ({start_time}-{end_time})")
        print(f"  Daily Hours: {daily_hours}, Days/Week: {days_per_week}")
        print(f"  Monthly Hours from Shift: {monthly_hours}")
        
        return monthly_hours
    
    def _get_shift_hours_per_day(self, employee_id: int) -> float:
        """Get daily hours from employee's assigned shift"""
        # Get the employee's active shift
        shift = self.db.query(Shift).filter(
            Shift.employee_id == employee_id,
            Shift.is_active == True
        ).first()
        
        if not shift:
            print(f"  No shift found for employee {employee_id}, using default 8 hours/day")
            return 8.0  # Default fallback
        
        # Calculate daily hours
        start_time = shift.start_time
        end_time = shift.end_time
        
        # Handle shifts that cross midnight (e.g., 16:00-00:00)
        if end_time < start_time:
            # Shift crosses midnight
            daily_hours = (24 - start_time.hour - start_time.minute/60) + (end_time.hour + end_time.minute/60)
        else:
            # Normal shift within same day
            daily_hours = (end_time.hour + end_time.minute/60) - (start_time.hour + start_time.minute/60)
        
        print(f"  Shift: {shift.shift_name} ({start_time}-{end_time}) = {daily_hours} hours/day")
        
        return daily_hours

    def get_attendance_data(self, start_date: Optional[date] = None, end_date: Optional[date] = None, 
                          employee_id: Optional[int] = None) -> List[dict]:
        """Get attendance data for reports"""
        query = self.db.query(Attendance).join(Employee)
        
        if start_date:
            query = query.filter(Attendance.date >= start_date.strftime('%Y-%m-%d'))
        if end_date:
            query = query.filter(Attendance.date <= end_date.strftime('%Y-%m-%d'))
        if employee_id:
            query = query.filter(Attendance.employee_id == employee_id)
            
        attendance_records = query.order_by(Attendance.date.desc()).all()
        
        report_data = []
        for record in attendance_records:
            report_data.append({
                'Date': record.date,
                'Employee ID': record.employee.employee_id,
                'Employee Name': record.employee.name,
                'Department': record.employee.department or 'N/A',
                'Check In': record.check_in.strftime('%H:%M:%S') if record.check_in else 'N/A',
                'Check Out': record.check_out.strftime('%H:%M:%S') if record.check_out else 'N/A',
                'Total Hours': f"{record.total_hours:.2f}" if record.total_hours else '0.00',
                'Status': record.status.title(),
                'Notes': record.notes or ''
            })
        
        return report_data

    def generate_pdf_report(self, start_date: Optional[date] = None, end_date: Optional[date] = None, 
                           employee_id: Optional[int] = None) -> io.BytesIO:
        """Generate PDF attendance report"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Build content
        content = []
        
        # Title
        title = "Attendance Report"
        if start_date and end_date:
            title += f" ({start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')})"
        content.append(Paragraph(title, title_style))
        content.append(Spacer(1, 20))
        
        # Get data
        data = self.get_attendance_data(start_date, end_date, employee_id)
        
        if not data:
            content.append(Paragraph("No attendance records found for the specified criteria.", styles['Normal']))
        else:
            # Create table
            table_data = [['Date', 'Employee ID', 'Name', 'Department', 'Check In', 'Check Out', 'Hours', 'Status']]
            
            for record in data:
                table_data.append([
                    record['Date'],
                    record['Employee ID'],
                    record['Employee Name'][:20] + '...' if len(record['Employee Name']) > 20 else record['Employee Name'],
                    record['Department'][:10] + '...' if len(record['Department']) > 10 else record['Department'],
                    record['Check In'],
                    record['Check Out'],
                    record['Total Hours'],
                    record['Status']
                ])
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            content.append(table)
            
            # Summary
            content.append(Spacer(1, 20))
            summary_text = f"Total Records: {len(data)}"
            content.append(Paragraph(summary_text, styles['Normal']))
        
        # Generate report info
        content.append(Spacer(1, 30))
        report_info = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        content.append(Paragraph(report_info, styles['Italic']))
        
        # Build PDF
        doc.build(content)
        buffer.seek(0)
        return buffer

    def generate_excel_report(self, start_date: Optional[date] = None, end_date: Optional[date] = None, 
                             employee_id: Optional[int] = None) -> io.BytesIO:
        """Generate Excel attendance report"""
        data = self.get_attendance_data(start_date, end_date, employee_id)
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Attendance Report', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Attendance Report']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary sheet
            summary_data = {
                'Metric': ['Total Records', 'Date Range', 'Generated On'],
                'Value': [
                    len(data),
                    f"{start_date or 'All'} to {end_date or 'All'}",
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        buffer.seek(0)
        return buffer

    def get_employee_summary(self, employee_id: int, start_date: Optional[date] = None, 
                           end_date: Optional[date] = None) -> dict:
        """Get summary statistics for an employee"""
        
        # If no date range provided, use a reasonable default (last 30 days)
        if not start_date or not end_date:
            end_date = date.today()
            start_date = end_date - timedelta(days=30)
        
        # Get attendance records for the period
        query = self.db.query(Attendance).filter(Attendance.employee_id == employee_id)
        query = query.filter(Attendance.date >= start_date.strftime('%Y-%m-%d'))
        query = query.filter(Attendance.date <= end_date.strftime('%Y-%m-%d'))
        records = query.all()
        
        # Calculate expected working days (Monday to Friday only)
        current_date = start_date
        expected_working_days = 0
        
        while current_date <= end_date:
            # Count weekdays only (Monday=0, Sunday=6)
            if current_date.weekday() < 5:  # Monday to Friday
                expected_working_days += 1
            current_date += timedelta(days=1)
        
        # Count actual attendance
        present_days = len([r for r in records if r.check_in is not None])
        late_days = len([r for r in records if r.check_in and r.check_in.time() > datetime.strptime("09:00", "%H:%M").time()])
        total_hours = sum([r.total_hours or 0 for r in records])
        
        # Calculate absent days correctly: Expected - Present
        absent_days = max(0, expected_working_days - present_days)
        
        # Calculate attendance rate: Present / Expected * 100
        attendance_rate = round((present_days / expected_working_days * 100) if expected_working_days > 0 else 0, 2)
        
        return {
            'total_days': expected_working_days,  # Expected working days in period
            'present_days': present_days,         # Days with check-in
            'late_days': late_days,              # Days late (after 9:00 AM)
            'absent_days': absent_days,          # Expected - Present
            'total_hours': round(total_hours, 2), # Sum of all hours worked
            'attendance_rate': attendance_rate    # (Present / Expected) * 100
        }

    def get_hourly_rate_analysis(self, start_date: Optional[date] = None, end_date: Optional[date] = None, 
                                employee_id: Optional[int] = None) -> List[dict]:
        """Calculate hourly rates for employees based on attendance and salary data"""
        
        # Get attendance data with total hours
        query = self.db.query(
            Employee.id,
            Employee.employee_id,
            Employee.name,
            Employee.department,
            Employee.salary_rate,
            func.sum(func.coalesce(Attendance.total_hours, 0)).label('total_hours')
        ).join(Attendance).filter(
            Employee.is_active == True,
            Attendance.total_hours.isnot(None)  # Only include records with actual hours
        )
        
        if start_date:
            query = query.filter(Attendance.date >= start_date.strftime('%Y-%m-%d'))
        if end_date:
            query = query.filter(Attendance.date <= end_date.strftime('%Y-%m-%d'))
        if employee_id:
            query = query.filter(Employee.id == employee_id)
            
        query = query.group_by(Employee.id, Employee.employee_id, Employee.name, 
                              Employee.department, Employee.salary_rate)
        
        results = query.all()
        
        # Debug: Print the SQL query result
        print(f"DEBUG: Hourly rate query results:")
        for result in results:
            print(f"  Employee: {result.name} (ID: {result.id})")
            print(f"  Total Hours from DB: {result.total_hours}")
            
            # Debug: Show individual attendance records for this employee
            individual_records = self.db.query(Attendance).filter(
                Attendance.employee_id == result.id
            )
            if start_date:
                individual_records = individual_records.filter(Attendance.date >= start_date.strftime('%Y-%m-%d'))
            if end_date:
                individual_records = individual_records.filter(Attendance.date <= end_date.strftime('%Y-%m-%d'))
            
            individual_records = individual_records.all()
            print(f"  Individual records:")
            for record in individual_records:
                print(f"    Date: {record.date}, Hours: {record.total_hours}")
            print(f"  Manual sum: {sum([r.total_hours or 0 for r in individual_records])}")
        
        hourly_data = []
        total_salary_cost = 0
        total_hours_worked = 0
        
        for result in results:
            # Debug: Print salary rate information
            print(f"DEBUG: Employee {result.name} ({result.employee_id})")
            print(f"  Salary Rate from DB: {result.salary_rate} cents/hour")
            
            # Calculate monthly salary based on Employee table salary_rate (primary source)
            if result.salary_rate > 0:
                # Convert cents per hour to AED per hour
                hourly_rate_aed = result.salary_rate / 1000  # Convert from cents to AED
                
                # Calculate monthly salary: Shift Hours × Days in Month × Hourly Rate
                shift_hours_per_day = self._get_shift_hours_per_day(result.id)
                days_in_month = 28  # Use 28 days as requested
                monthly_hours = shift_hours_per_day * days_in_month
                monthly_salary = monthly_hours * hourly_rate_aed
                
                print(f"  Salary Rate: {result.salary_rate} cents = {hourly_rate_aed} AED/hour")
                print(f"  Shift Hours per Day: {shift_hours_per_day}")
                print(f"  Monthly Calculation: {shift_hours_per_day} hours/day × {days_in_month} days × {hourly_rate_aed} AED/hr = {monthly_salary} AED")
            else:
                # Fallback: Get latest salary record for this employee
                latest_salary = self.db.query(Salary).filter(
                    Salary.employee_id == result.id
                ).order_by(Salary.created_at.desc()).first()
                
                if latest_salary and latest_salary.gross_salary > 0:
                    monthly_salary = latest_salary.gross_salary
                else:
                    # Default monthly salary if no data available
                    monthly_salary = 2000.0  # Default 2000 AED
            
            total_hours = float(result.total_hours or 0)
            
            # Calculate earned salary directly: Hourly Rate × Actual Hours Worked
            if total_hours > 0:
                # Use the actual hourly rate from employee table (not calculated from monthly)
                actual_hourly_rate = hourly_rate_aed if result.salary_rate > 0 else 0
                per_minute_rate = actual_hourly_rate / 60
                # Direct calculation: Hourly Rate × Total Hours Worked
                earned_salary = actual_hourly_rate * total_hours
                
                print(f"  Direct Calculation:")
                print(f"    Hourly Rate: {actual_hourly_rate} AED/hour")
                print(f"    Total Hours: {total_hours}")
                print(f"    Earned Salary: {actual_hourly_rate} × {total_hours} = {earned_salary} AED")
                
                # For display purposes, show the actual hourly rate (not monthly/hours)
                hourly_rate = actual_hourly_rate
            else:
                hourly_rate = hourly_rate_aed if result.salary_rate > 0 else 0
                per_minute_rate = hourly_rate / 60 if hourly_rate > 0 else 0
                earned_salary = 0
            
            employee_data = {
                'employee_id': result.employee_id,
                'employee_name': result.name,
                'department': result.department or 'N/A',
                'monthly_salary': round(monthly_salary, 2),
                'total_hours': round(total_hours, 2),
                'hourly_rate': round(hourly_rate, 2),
                'per_minute_rate': round(per_minute_rate, 3),
                'earned_salary': round(earned_salary, 2),
                'salary_difference': round(earned_salary - monthly_salary, 2),
                'cost_efficiency': round(hourly_rate, 2) if hourly_rate > 0 else 0
            }
            
            hourly_data.append(employee_data)
            total_salary_cost += earned_salary  # Use earned salary based on actual hours worked
            total_hours_worked += total_hours
        
        # Calculate summary statistics
        if hourly_data:
            avg_hourly_rate = sum(emp['hourly_rate'] for emp in hourly_data) / len(hourly_data)
            avg_per_minute_rate = avg_hourly_rate / 60
            total_earned_salary = sum(emp['earned_salary'] for emp in hourly_data)
            total_salary_difference = sum(emp['salary_difference'] for emp in hourly_data)
        else:
            avg_hourly_rate = 0
            avg_per_minute_rate = 0
            total_earned_salary = 0
            total_salary_difference = 0
        
        # Add summary data
        summary = {
            'total_employees': len(hourly_data),
            'total_salary_cost': round(total_salary_cost, 2),
            'total_earned_salary': round(total_earned_salary, 2),
            'total_salary_difference': round(total_salary_difference, 2),
            'total_hours_worked': round(total_hours_worked, 2),
            'average_hourly_rate': round(avg_hourly_rate, 2),
            'average_per_minute_rate': round(avg_per_minute_rate, 3),
            'period_start': start_date.strftime('%Y-%m-%d') if start_date else 'N/A',
            'period_end': end_date.strftime('%Y-%m-%d') if end_date else 'N/A'
        }
        
        return {
            'employees': hourly_data,
            'summary': summary
        }

    def generate_pdf_report_with_rates(self, start_date: Optional[date] = None, end_date: Optional[date] = None, 
                                      employee_id: Optional[int] = None) -> io.BytesIO:
        """Generate PDF report with attendance data and hourly rate analysis"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        period_text = ""
        if start_date and end_date:
            period_text = f" ({start_date} to {end_date})"
        
        story.append(Paragraph(f"Attendance Report with Hourly Rate Analysis{period_text}", title_style))
        story.append(Spacer(1, 20))
        
        # Get regular attendance data
        attendance_data = self.get_attendance_data(start_date, end_date, employee_id)
        
        if attendance_data:
            # Attendance Table
            story.append(Paragraph("Attendance Records", styles['Heading2']))
            story.append(Spacer(1, 10))
            
            # Create attendance table (limit to first 10 records for space)
            attendance_headers = ['Date', 'Employee', 'Check In', 'Check Out', 'Hours', 'Status']
            attendance_table_data = [attendance_headers]
            
            for record in attendance_data[:10]:  # Limit for space
                attendance_table_data.append([
                    str(record['Date']),
                    record['Employee Name'],
                    record['Check In'],
                    record['Check Out'],
                    record['Total Hours'],
                    record['Status']
                ])
            
            attendance_table = Table(attendance_table_data)
            attendance_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(attendance_table)
            story.append(Spacer(1, 30))
        
        # New Page for Hourly Rate Analysis
        from reportlab.platypus import PageBreak
        story.append(PageBreak())
        
        # Hourly Rate Analysis
        story.append(Paragraph("Hourly Rate Analysis", styles['Heading1']))
        story.append(Spacer(1, 20))
        
        # Get hourly rate data
        hourly_data = self.get_hourly_rate_analysis(start_date, end_date, employee_id)
        
        if hourly_data['employees']:
            # Summary Box
            summary = hourly_data['summary']
            summary_text = f"""
            <b>Summary Statistics:</b><br/>
            • Total Employees: {summary['total_employees']}<br/>
            • Total Salary Cost: {summary['total_salary_cost']} AED<br/>
            • Total Hours Worked: {summary['total_hours_worked']} hours<br/>
            • Average Hourly Rate: {summary['average_hourly_rate']} AED/hour<br/>
            • Average Per Minute Rate: {summary['average_per_minute_rate']} AED/minute<br/>
            • Period: {summary['period_start']} to {summary['period_end']}
            """
            
            story.append(Paragraph(summary_text, styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Hourly Rate Table
            story.append(Paragraph("Employee Hourly Rate Breakdown", styles['Heading2']))
            story.append(Spacer(1, 10))
            
            rate_headers = ['Employee ID', 'Name', 'Department', 'Monthly Salary (AED)', 
                           'Total Hours', 'Hourly Rate (AED)', 'Per Minute (AED)']
            rate_table_data = [rate_headers]
            
            for emp in hourly_data['employees']:
                rate_table_data.append([
                    emp['employee_id'],
                    emp['employee_name'],
                    emp['department'],
                    f"{emp['monthly_salary']:.2f}",
                    f"{emp['total_hours']:.2f}",
                    f"{emp['hourly_rate']:.2f}",
                    f"{emp['per_minute_rate']:.3f}"
                ])
            
            rate_table = Table(rate_table_data)
            rate_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),  # Right align numbers
            ]))
            
            story.append(rate_table)
            story.append(Spacer(1, 20))
            
            # Calculation Example
            if hourly_data['employees']:
                example_emp = hourly_data['employees'][0]
                example_text = f"""
                <b>Calculation Example ({example_emp['employee_name']}):</b><br/>
                1. Find the hourly rate:<br/>
                   {example_emp['monthly_salary']} AED ÷ {example_emp['total_hours']} hours = {example_emp['hourly_rate']} AED/hour<br/><br/>
                2. Convert to per minute rate:<br/>
                   Since there are 60 minutes in an hour:<br/>
                   {example_emp['hourly_rate']} ÷ 60 = {example_emp['per_minute_rate']} AED/minute<br/><br/>
                ✅ <b>Per minute rate = {example_emp['per_minute_rate']} AED/minute</b>
                """
                
                story.append(Paragraph(example_text, styles['Normal']))
        else:
            story.append(Paragraph("No hourly rate data available for the selected period.", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer

    def generate_excel_report_with_rates(self, start_date: Optional[date] = None, end_date: Optional[date] = None, 
                                        employee_id: Optional[int] = None) -> io.BytesIO:
        """Generate Excel report with attendance data and hourly rate analysis"""
        buffer = io.BytesIO()
        
        # Get data
        attendance_data = self.get_attendance_data(start_date, end_date, employee_id)
        hourly_data = self.get_hourly_rate_analysis(start_date, end_date, employee_id)
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Sheet 1: Attendance Data
            if attendance_data:
                attendance_df = pd.DataFrame(attendance_data)
                attendance_df.to_excel(writer, sheet_name='Attendance Records', index=False)
            
            # Sheet 2: Hourly Rate Analysis
            if hourly_data['employees']:
                # Employee data
                hourly_df = pd.DataFrame(hourly_data['employees'])
                hourly_df.to_excel(writer, sheet_name='Hourly Rate Analysis', index=False, startrow=0)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Hourly Rate Analysis']
                
                # Add summary section
                summary = hourly_data['summary']
                summary_start_row = len(hourly_df) + 3
                
                # Summary headers and data
                summary_data = [
                    ['SUMMARY STATISTICS', ''],
                    ['Total Employees', summary['total_employees']],
                    ['Total Salary Cost (AED)', summary['total_salary_cost']],
                    ['Total Hours Worked', summary['total_hours_worked']],
                    ['Average Hourly Rate (AED)', summary['average_hourly_rate']],
                    ['Average Per Minute Rate (AED)', summary['average_per_minute_rate']],
                    ['Period Start', summary['period_start']],
                    ['Period End', summary['period_end']],
                    ['', ''],
                    ['CALCULATION EXAMPLE', ''],
                ]
                
                # Add calculation example if we have employees
                if hourly_data['employees']:
                    example_emp = hourly_data['employees'][0]
                    calculation_example = [
                        [f"Example: {example_emp['employee_name']}", ''],
                        ['1. Monthly Salary ÷ Total Hours = Hourly Rate', ''],
                        [f"{example_emp['monthly_salary']} AED ÷ {example_emp['total_hours']} hours", f"{example_emp['hourly_rate']} AED/hour"],
                        ['2. Hourly Rate ÷ 60 = Per Minute Rate', ''],
                        [f"{example_emp['hourly_rate']} ÷ 60", f"{example_emp['per_minute_rate']} AED/minute"],
                        ['✅ Per minute rate', f"{example_emp['per_minute_rate']} AED/minute"]
                    ]
                    summary_data.extend(calculation_example)
                
                # Write summary data
                for i, (label, value) in enumerate(summary_data):
                    worksheet.cell(row=summary_start_row + i + 1, column=1, value=label)
                    worksheet.cell(row=summary_start_row + i + 1, column=2, value=value)
                
                # Format the worksheet
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Header formatting
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                # Format headers
                for col in range(1, len(hourly_df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Format summary section
                summary_font = Font(bold=True)
                summary_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                
                # Format summary headers
                worksheet.cell(row=summary_start_row + 1, column=1).font = Font(bold=True, size=12)
                worksheet.cell(row=summary_start_row + 1, column=1).fill = summary_fill
                worksheet.cell(row=summary_start_row + 10, column=1).font = Font(bold=True, size=12)
                worksheet.cell(row=summary_start_row + 10, column=1).fill = summary_fill
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 3: Rate Comparison (if multiple employees)
            if len(hourly_data['employees']) > 1:
                comparison_data = []
                for emp in hourly_data['employees']:
                    comparison_data.append({
                        'Employee': emp['employee_name'],
                        'Hourly Rate (AED)': emp['hourly_rate'],
                        'Per Minute Rate (AED)': emp['per_minute_rate'],
                        'Efficiency Rank': 0  # Will be calculated
                    })
                
                # Sort by hourly rate and add ranking
                comparison_data.sort(key=lambda x: x['Hourly Rate (AED)'])
                for i, emp in enumerate(comparison_data):
                    emp['Efficiency Rank'] = i + 1
                
                comparison_df = pd.DataFrame(comparison_data)
                comparison_df.to_excel(writer, sheet_name='Rate Comparison', index=False)
        
        buffer.seek(0)
        return buffer
